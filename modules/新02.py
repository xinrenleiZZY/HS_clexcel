import os
import re
from openpyxl import load_workbook


def replace_excel_content(input_file, output_file):
    """
    专门用于替换Excel文件中的指定内容

    参数:
        input_file: 输入Excel文件路径（如上下班打卡_7月报_processed.xlsx）
        output_file: 输出Excel文件路径，默认为在输入文件名后加"_replaced"
    """
    try:
        # 自动生成输出文件名
        output_file = output_file

        # 打开Excel文件
        wb = load_workbook(input_file)

        # 替换模式列表（非单元格匹配，按优先级排序）
        patterns_to_replace = [
            # 1. 带分号的缺卡格式（增强匹配）
            r'正常（未排班）',
            r'缺卡\([^)]*\);',  # 匹配"缺卡(任意内容);"
            r'缺卡\(.*?\);',  # 备用模式，确保匹配
            # 2. 不带分号的缺卡格式
            r'缺卡\([^)]*\)',
            r'缺卡\(.*?\)',
            # 3. 补卡申请格式
            r'补卡申请（[^）]*）',
            r'补卡申请（.*?）',
            # 4. 正常(补卡)格式
            r'正常\(补卡\)-',
            # 5. 正常格式
            r'正常-',
            # 6. 双横线格式（多种可能的横线）
            r'--',
            r'— —',  # 全角横线
            r'——',  # 破折号
            # 7. 单独的缺卡
            r'缺卡',
            # 8. 各种换行符和空白字符
            r'\r\n|\r|\n|\t',
            # 9. 空格（多个连续空格）
            r' +',
            r'地点异常.*?;',
            r'\(补卡\)-',
            r'正常\(管理员校准、补卡\)-',
            r'正常\(休息\)',
            r'正常（休息）'
            r'正常\(管理员校准\)-',
            r'迟到\s*[\d.]*\s*分钟-?;',
            r'早退\s*[\d.]*\s*分钟-?;',
            r'旷工\s*[\d.]*\s*分钟-?;',
        ]
        patterns_to_replace2 = [
            r'迟到\s*[\d.]*\s*分钟-?',
            r'早退\s*[\d.]*\s*分钟-?',
            r'旷工\s*[\d.]*\s*分钟-?',
        ]

        # 处理每个工作表
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]

            # 只处理可见工作表
            if ws.sheet_state != 'visible':
                print(f"工作表 {sheet_name} 是隐藏的，已跳过")
                continue

            # 记录替换数量
            replace_count = 0

            # 遍历所有单元格进行替换
            for row in range(1, ws.max_row + 1):
                for col in range(1, ws.max_column + 1):
                    # 跳过姓名、员工ID、部门（假设这些列是前3列）
                    if col <= 3:
                        continue

                    cell = ws.cell(row=row, column=col)
                    if cell.value is not None:
                        original_value = str(cell.value)
                        cell_text = original_value

                        # 应用所有替换模式
                        for pattern in patterns_to_replace:
                            cell_text = re.sub(pattern, '', cell_text)
                        # 应用所有替换模式
                        for pattern in patterns_to_replace2:
                            cell_text = re.sub(pattern, ';', cell_text)

                        # 最终清理
                        cell_text = cell_text.strip()

                        # 如果内容有变化，更新单元格并计数
                        if cell_text != original_value:
                            cell.value = cell_text if cell_text else ""
                            replace_count += 1

            print(f"工作表 {sheet_name} 完成替换，共处理 {replace_count} 个单元格")

        # 保存处理后的文件
        wb.save(output_file)
        wb.close()

        print(f"替换完成，已保存至: {output_file}")
        return output_file

    except Exception as e:
        print(f"替换过程出错: {str(e)}")
        return None


if __name__ == "__main__":
    import sys

    if len(sys.argv) == 3:
        input_file = sys.argv[1]
        output_file = sys.argv[2] if len(sys.argv) > 2 else None
        replace_excel_content(input_file, output_file)
    else:
        print("用法: python 1.py <输入文件> [输出文件]")

